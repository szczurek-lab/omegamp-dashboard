#!/usr/bin/env python3
"""Build the manuscript supplementary data deposits.

Run from the repository root:
    python scripts/build_supplementary.py

Outputs (overwritten on every run) go to supplementary/ :

  Data_S1_prototype_set.xlsx    500 active + 500 inactive DBAASP prototypes,
                                one sheet each                       (Data S1)
  Data_S2_wetlab_dataset.xlsx   one sheet per measurement type       (Data S2)

Inputs (read-only):
    data/prototypes/curated-AMPs_samples.fasta       - Data S1 active set
    data/prototypes/curated-Non-AMPs_samples.fasta   - Data S1 inactive set
    data/omegamp_reference_table.csv                 - peptide metadata
    data/*.csv                                       - processed assay tables
                                                       (written by build_data.py)

Data S2 covers the 215 wet-lab characterised OmegAMP peptides plus the two
antibiotic controls (Polymyxin B, Levofloxacin) wherever they were measured
alongside.  Peptides present in the raw assay files but outside this set
(previously published comparator peptides run on the same plates) are dropped
and reported on stdout, so the deposit matches the manuscript's n = 215.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DATA = Path("data")
PROTO = DATA / "prototypes"
OUT = Path("supplementary")

S1_XLSX = "Data_S1_prototype_set.xlsx"
XLSX = "Data_S2_wetlab_dataset.xlsx"

# The pipeline tracks two prototypes by a family shorthand ("cecropin",
# "sarcotoxin") that is not the peptide's name -- the peptides actually tested
# are cecropin P1 (SWLSKTAKKLENSAKKRISEGIAIAIQGGPR) and sarcotoxin IA
# (GWLKKIGKKIERVGQHTRDATIQGLGIAQQAANVAATAR).  Every spelling that reaches the
# deposit is mapped to the correct name, in short_name, prototype and family.
NAME_ALIASES = {
    "cecropin":      "Cecropin-P1",
    "cecropin-P1":   "Cecropin-P1",
    "sarcotoxin":    "sarcotoxin-1A",
}

# The reference table labels de novo modes without the D of their short names.
DENOVO_MODES = {"OmegAMP-P": "OmegAMP-DP", "OmegAMP-T": "OmegAMP-DT"}

# Sampling mode -> conditioning strategy, per scripts/reproduce/gen_figure3_flavors.sh:
# D de novo, A analog of a prototype sequence, M motif-guided,
# T property-targeted (length / charge / hydrophobicity), P prototype-derived
# conditioning, U unconditional.
CONDITIONING = {
    "OmegAMP-DP":  "de novo, prototype-derived conditioning",
    "OmegAMP-DT":  "de novo, property-targeted (length, charge, hydrophobicity)",
    "OmegAMP-AU":  "analog of a prototype, unconditional",
    "OmegAMP-AT":  "analog of a prototype, property-targeted",
    "OmegAMP-AP":  "analog of a prototype, prototype-derived conditioning",
    "OmegAMP-AMT": "analog of a prototype, motif-guided and property-targeted",
    "OmegAMP-MT":  "motif-guided, property-targeted",
}
NOT_GENERATED = {
    "prototype": "not generated (parent prototype peptide)",
    "control":   "not generated (antibiotic control)",
}

# Strain both membrane assays are run against; its MIC is the at-MIC dose.
MOA_STRAIN = "A. baumannii ATCC 19606 (-)"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def read_fasta(path):
    """Yield (id, sequence) pairs from a two-line-per-record FASTA."""
    ident, seq = None, []
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith(">"):
            if ident is not None:
                yield ident, "".join(seq)
            ident, seq = line[1:].strip(), []
        else:
            seq.append(line)
    if ident is not None:
        yield ident, "".join(seq)


def load(name):
    """Read a processed table from data/, applying correct prototype names."""
    df = pd.read_csv(DATA / f"{name}.csv")
    for col in ("short_name", "prototype", "prototype_display", "family"):
        if col in df.columns:
            df[col] = df[col].replace(NAME_ALIASES)
    return df


def restrict(df, keep, label):
    """Keep only peptides in the deposited set; report what was dropped."""
    dropped = sorted(set(df["short_name"]) - keep)
    if dropped:
        print(f"  {label}: dropped {len(dropped)} non-deposit peptides "
              f"({', '.join(dropped[:4])}{', ...' if len(dropped) > 4 else ''})")
    return df[df["short_name"].isin(keep)].reset_index(drop=True)


def order_by(df, order):
    """Sort rows by the reference-table peptide order, stably."""
    rank = {n: i for i, n in enumerate(order)}
    return (df.assign(_r=df["short_name"].map(rank))
              .sort_values("_r", kind="stable")
              .drop(columns="_r")
              .reset_index(drop=True))


# ─────────────────────────────────────────────────────────────────────────────
# Data S1 - in silico prototype set
# ─────────────────────────────────────────────────────────────────────────────

def build_data_s1():
    """One workbook, one sheet each for the active and inactive prototypes."""
    sheets = []
    for fasta, sheet in (
        (PROTO / "curated-AMPs_samples.fasta", "Active prototypes"),
        (PROTO / "curated-Non-AMPs_samples.fasta", "Inactive prototypes"),
    ):
        if not fasta.exists():
            sys.exit(f"ERROR: missing prototype FASTA {fasta}")
        rows = [{"prototype_id": ident, "sequence": seq}
                for ident, seq in read_fasta(fasta)]
        sheets.append((sheet, pd.DataFrame(rows)))

    path = OUT / S1_XLSX
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets:
            df.to_excel(writer, sheet_name=sheet, index=False)
        style_workbook(writer.book)

    print(f"{S1_XLSX}: " + ", ".join(f"{len(df)} {s.lower()}" for s, df in sheets))


# ─────────────────────────────────────────────────────────────────────────────
# Data S2 - wet-lab workbook
# ─────────────────────────────────────────────────────────────────────────────

def membrane_sheet(assay, keep, order, at_mic_conc):
    """Merge the at-MIC and fixed-concentration runs of NPN / DiSC3(5).

    Both assays use A. baumannii ATCC 19606, so the at-MIC concentration is
    that peptide's MIC against the same strain; it stays blank where the MIC
    is right-censored (>64) or the agent has no MIC (Triton X-100).
    """
    at_mic = load(assay).assign(condition="at MIC")
    at_mic["concentration_uM"] = at_mic["short_name"].map(at_mic_conc)
    fixed = load(f"{assay}_fc").assign(condition="fixed concentration")
    df = pd.concat([at_mic, fixed], ignore_index=True)
    df = df[["short_name", "condition", "concentration_uM", "MaxRel", "AUC"]]
    df = df.rename(columns={"MaxRel": "MaxRel_percent"})
    # Triton X-100 is the complete-lysis positive control of both assays.
    df = restrict(df, keep | {"Triton"}, assay.upper())
    return order_by(df, order).sort_values(
        "condition", kind="stable").reset_index(drop=True)


def murine_sheet(model, keep_all):
    """Per-mouse in vivo readouts, with units made explicit."""
    df = load(f"murine_{model}")
    df = df.rename(columns={"replicate": "mouse", "value": "value"})
    df["assay"] = df["assay"].map({"cfu": "bacterial burden",
                                   "weight": "body weight"})
    df["unit"] = df["assay"].map({"bacterial burden": "CFU per g tissue",
                                  "body weight": "% of day 0"})
    dropped = sorted(set(df["short_name"]) - keep_all)
    if dropped:
        print(f"  murine {model}: unmapped groups {dropped}")
    return df[["short_name", "assay", "day", "mouse", "value", "unit"]]


def build_data_s2():
    ref = load("omegamp_reference_table")
    order = ref["short_name"].tolist()
    keep = set(order)                       # 215 peptides + 2 antibiotic controls
    keep_murine = keep | {"Control"}        # untreated infection control

    n_pep = int((ref["category"] != "control").sum())
    print(f"Data S2: {n_pep} peptides + "
          f"{len(ref) - n_pep} antibiotic controls")

    # De novo modes are stored without the D that their short names carry
    # (Ω-DP-1 is OmegAMP-P in the reference table); restore it here so the
    # workbook uses the same OmegAMP-D* / OmegAMP-A* labels as Figure 3.
    meta = ref.copy()
    meta["generation_mode"] = meta["generation_mode"].replace(DENOVO_MODES)
    meta["conditioning_strategy"] = meta["generation_mode"].map(CONDITIONING)
    meta.loc[meta["conditioning_strategy"].isna(), "conditioning_strategy"] = (
        meta["category"].map(NOT_GENERATED))
    unlabelled = meta["conditioning_strategy"].isna().sum()
    if unlabelled:
        print(f"  WARNING: {unlabelled} peptides without a conditioning label")
    # prototype_display carries the literature name of prototypes that the
    # pipeline tracks by DBAASP accession (DBAASPS_20015 -> As-CATH4-6L).
    meta["prototype"] = meta["prototype_display"]

    # Physicochemical descriptors are omitted deliberately: they are derivable
    # from the sequence, unlike the measured values this deposit exists for.
    metadata = meta[["short_name", "sequence", "generation_mode",
                     "conditioning_strategy", "prototype", "objective",
                     "round"]]

    mic = restrict(load("mic"), keep, "MIC")
    at_mic_conc = dict(zip(mic["short_name"],
                           pd.to_numeric(mic[MOA_STRAIN], errors="coerce")))

    cc50 = restrict(load("cc50"), keep, "CC50")
    hc50 = restrict(load("hc50"), keep, "HC50")
    bestsel = restrict(load("bestsel"), keep, "BeStSel")

    lps = restrict(load("lps_binding"), keep, "LPS binding").rename(columns={
        "concentration": "concentration_uM",
        "calcium": "calcium_added",
        "BC_displacement": "BC_displacement_percent"})

    dna = restrict(load("dna_binding"), keep, "DNA CD").rename(columns={
        "concentration": "concentration_uM"})
    dna["unit"] = dna["metric"].map(
        {"deltaA": "mdeg", "deltaAmax": "mdeg",
         "deltaLambdaNeg": "nm", "deltaLambdaPos": "nm"})

    prot = restrict(load("proteolysis"), keep, "Proteolysis").rename(columns={
        "rep1": "replicate_1_percent",
        "rep2": "replicate_2_percent",
        "mean": "mean_percent_remaining"})

    novelty = restrict(load("denovo_training_identity"), keep, "Novelty").rename(
        columns={"best_id_lenmatched": "max_identity_to_training_set",
                 "best_hit_sequence": "closest_training_sequence"})
    novelty = novelty.drop(columns=["length"])

    sheets = [
        ("Peptide metadata", order_by(metadata, order),
         "Identity and design provenance of every deposited peptide. "
         "generation_mode is the OmegAMP sampling mode (D de novo, A analog "
         "of a prototype, M motif-guided, T property-targeted, P "
         "prototype-derived conditioning, U unconditional), spelled out in "
         "conditioning_strategy; prototype names the parent peptide for "
         "analogs and motif designs; objective is the design goal of the "
         "series and round the design iteration. Physicochemical descriptors "
         "are not deposited -- they follow from the sequence. Polymyxin B and "
         "levofloxacin are non-ribosomal antibiotics, so their sequence cell "
         "is blank."),
        ("MIC", order_by(mic, order),
         "Minimum inhibitory concentration (umol L-1) against the 20-strain "
         "panel, one column per strain. Values are the modal MIC of "
         "independent replicates; '>64' means no inhibition at the highest "
         "concentration tested."),
        ("CC50", order_by(cc50, order),
         "Cytotoxicity against HEK293T cells (umol L-1), from non-linear "
         "regression of the dose-response curve. '>128' is right-censored: "
         "the value exceeds the tested range."),
        ("HC50", order_by(hc50, order),
         "Haemolysis of human erythrocytes (umol L-1), from non-linear "
         "regression of the dose-response curve. '>128' is right-censored."),
        ("NPN", membrane_sheet("npn", keep, order, at_mic_conc),
         "Outer-membrane permeabilisation of A. baumannii ATCC 19606 by NPN "
         "uptake. MaxRel is peak fluorescence relative to untreated baseline "
         "(%); AUC integrates the fluorescence-time curve. Measured at each "
         "peptide's own MIC against this strain, and at a family-anchored "
         "fixed concentration for the analog series; concentration_uM gives "
         "the dose in both regimes and is blank where the MIC is censored "
         "(>64) or the agent has no MIC (Triton X-100)."),
        ("DiSC3(5)", membrane_sheet("disc", keep, order, at_mic_conc),
         "Cytoplasmic-membrane depolarisation of A. baumannii ATCC 19606 by "
         "DiSC3(5) release, same metrics and two concentration regimes as the "
         "NPN sheet."),
        ("LPS binding", order_by(lps, order),
         "Displacement of BODIPY TR cadaverine from LPS (%), three "
         "independent replicates. Without added Ca2+ the peptide was tested "
         "at 2, 8 and 32 umol L-1; the Ca2+ charge-shielding condition "
         "(calcium_added = yes) was run at 32 umol L-1 only. family is the "
         "prototype series (one assay plate per series); Polymyxin B is the "
         "positive control and was run on all four plates, hence four sets "
         "of replicates."),
        ("DNA CD", order_by(dna, order),
         "Circular-dichroism perturbation of DNA by the bZIP motif series. "
         "deltaA is the ellipticity change (mdeg) at 12.5, 25 and 50 umol "
         "L-1 peptide and deltaAmax its maximum over those concentrations "
         "(concentration_uM therefore blank); deltaLambdaPos / "
         "deltaLambdaNeg are shifts (nm) of the positive and negative CD "
         "bands at each concentration."),
        ("BeStSel", order_by(bestsel, order),
         "Secondary-structure composition (%) from BeStSel deconvolution of "
         "peptide CD spectra in four solvent conditions: H2O, TFE/H2O 3:2 "
         "(v/v), 10 mM SDS/H2O, MeOH/H2O 1:1 (v/v). fH helix, f_beta_anti / "
         "f_beta_par antiparallel and parallel sheet, fturn turn, fothers "
         "unordered."),
        ("Murine skin", murine_sheet("skin", keep_murine),
         "Skin-scarification model, A. baumannii ATCC 19606, CD-1 mice. "
         "Peptides applied topically at 10x MIC 1 h post-infection. Bacterial "
         "burden per excised wound at days 2 and 4 (CFU per g) and body "
         "weight relative to day 0 (%), one row per mouse. Twelve mice per "
         "group are weighed until the day-2 sacrifice and the surviving six "
         "thereafter, so each CFU timepoint has n = 6. Control is untreated "
         "infection; Polymyxin B is the antibiotic benchmark."),
        ("Murine thigh", murine_sheet("thigh", keep_murine),
         "Neutropenic thigh model, A. baumannii ATCC 19606. Peptides dosed "
         "intraperitoneally at 10x MIC 2 h post-infection; burden at days 6 "
         "and 8 (CFU per g) and body weight relative to day 0 (%), one row "
         "per mouse. Twelve mice per group are weighed through day 6 and the "
         "surviving six on day 8, so each CFU timepoint has n = 6. "
         "Levofloxacin is the antibiotic control."),
        ("Proteolytic stability", order_by(prot, order),
         "Intact peptide remaining (%) over 6 h of protease exposure, two "
         "independent replicates and their mean, as reported by the assay "
         "(replicate 1 normalised to 100% at t = 0). Used to select the leads "
         "advanced to the systemic thigh model."),
        ("De novo novelty", order_by(novelty, order),
         "Maximum sequence identity of each de novo peptide to the "
         "generative training set (length-matched alignment), with the "
         "closest training sequence."),
    ]

    path = OUT / XLSX
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        readme = pd.DataFrame(
            [{"Sheet": name, "Rows": len(df), "Contents": doc}
             for name, df, doc in sheets])
        readme.to_excel(writer, sheet_name="README", index=False)
        for name, df, _ in sheets:
            df.to_excel(writer, sheet_name=name, index=False)
        style_workbook(writer.book)

    print(f"{XLSX}: {len(sheets) + 1} sheets")
    for name, df, _ in sheets:
        print(f"  {name:<22} {len(df):>4} rows x {len(df.columns)} cols")


def style_workbook(book):
    """Bold frozen header row, readable column widths, wrapped README text."""
    for ws in book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for i, col in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 400)), 1):
            longest = max((len(str(c.value)) for c in col if c.value is not None),
                          default=8)
            ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 46)
        if ws.title == "README":
            ws.column_dimensions["C"].width = 100
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                row[0].alignment = Alignment(vertical="top", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    missing = [f for f in ("omegamp_reference_table.csv", "mic.csv", "murine_skin.csv")
               if not (DATA / f).exists()]
    if missing:
        sys.exit(f"ERROR: missing input(s) in {DATA}/: {', '.join(missing)}\n"
                 "Run python scripts/build_data.py first.")

    OUT.mkdir(exist_ok=True)
    build_data_s1()
    build_data_s2()
    print(f"\nWritten to {OUT.resolve()}")
